# -*- coding: utf-8 -*-
"""
把旧模板(副本列表字段整理模板.xlsx)的岗位数据迁移到新模板(岗位导入模板.xlsx)
旧表头: 岗位名称 | 公司行业 | 岗位分类 | 地区 | 招聘周期 | 发布时间 | 截止时间 | 学员数
新表头: 岗位分类* | 岗位名称* | 公司名称* | 公司类别* | 城市* | 招聘周期* | 项目时间* | 截止时间 | 部门 | 岗位链接 | 公司官网
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import re

# 读旧数据
old_wb = openpyxl.load_workbook("/Users/hw/workspace/OSGPrj/docs/副本列表字段整理模板.xlsx", data_only=True)
old_ws = old_wb["岗位列表"]

# 旧表头列映射 (0-indexed)
# R1: 岗位名称(0) | 公司行业(1) | 岗位分类(2) | 地区(3) | 招聘周期(4) | 发布时间(5) | 截止时间(6) | 学员数(7)

# 岗位分类映射: 旧值 -> 新 position_category
def map_category(raw):
    raw = (raw or "").strip()
    low = raw.lower()
    if "sa" in low or "summer" in low:
        return "summer"
    if "ft" in low or "full" in low or "full-time" in low:
        return "fulltime"
    if "off" in low or "off-cycle" in low:
        return "offcycle"
    if "spring" in low:
        return "spring"
    if "event" in low or "insight" in low:
        return "events"
    if "intern" in low:
        return "summer"  # consulting intern -> summer
    if "swe" in low or "pm" in low:
        return "summer"
    return "summer"

# 公司行业 -> 公司类别映射
def map_company_type(industry):
    industry = (industry or "").strip()
    mapping = {
        "Consulting": "Consulting",
        "Elite Boutique": "Investment Bank",
        "Bulge Bracket": "Investment Bank",
        "Middle Market": "Investment Bank",
        "Buyside": "PE/VC",
    }
    return mapping.get(industry, "Other")

# 从岗位分类中提取项目时间
def extract_project_year(category_raw, position_name):
    text = (category_raw or "") + " " + (position_name or "")
    years = re.findall(r"(20\d{2})", text)
    if years:
        return years[0]
    return "2026"

# 从地区推断城市
def normalize_city(raw):
    raw = (raw or "").strip()
    if not raw:
        return ""
    # 特殊处理
    aliases = {
        "US": "New York",
        "Event": "",
        "Handshake": "",
        "it is a 2027SA": "",
        "Hong Kong SAR, Singapore": "Hong Kong",
        "Hong Kong SAR": "Hong Kong",
        "St Louis, MO": "St Louis",
    }
    if raw in aliases:
        return aliases[raw]
    return raw

# 截止时间格式化
def format_deadline(val):
    if val is None:
        return ""
    s = str(val).strip()
    if "Rolling" in s or "ASAP" in s:
        return "Rolling ASAP"
    # datetime 对象
    if hasattr(val, 'strftime'):
        return val.strftime("%Y-%m-%d")
    return s

# 读旧数据行
old_rows = []
for row_idx in range(2, old_ws.max_row + 1):
    cells = [old_ws.cell(row=row_idx, column=c).value for c in range(1, 9)]
    if not cells[0]:  # 岗位名称为空跳过
        continue
    old_rows.append(cells)

print(f"读取到 {len(old_rows)} 条旧数据")

# 打开新模板
new_wb = openpyxl.load_workbook("/Users/hw/workspace/OSGPrj/docs/岗位导入模板.xlsx")
new_ws = new_wb["岗位列表"]

# 删除示例数据 (R3-R5)
for row_idx in range(5, 2, -1):
    new_ws.delete_rows(row_idx)

# 写入转换后的数据
data_font = Font(name="微软雅黑", size=10)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
data_align = Alignment(horizontal="left", vertical="center")

for i, old in enumerate(old_rows):
    row_num = i + 3  # 从第3行开始(第1行表头，第2行说明)
    position_name = old[0] or ""
    industry = old[1] or ""
    category_raw = old[2] or ""
    city_raw = old[3] or ""
    cycle = old[4] or ""
    # publish_time = old[5]  # 不需要，系统自动
    deadline = old[6]
    # student_count = old[7]  # 不需要

    new_row = [
        map_category(category_raw),           # 岗位分类
        position_name,                        # 岗位名称
        "",                                   # 公司名称 (旧数据没有)
        map_company_type(industry),           # 公司类别
        normalize_city(city_raw),             # 城市
        cycle,                                # 招聘周期
        extract_project_year(category_raw, position_name),  # 项目时间
        format_deadline(deadline),            # 截止时间
        "",                                   # 部门
        "",                                   # 岗位链接
        "",                                   # 公司官网
    ]

    for col_idx, value in enumerate(new_row, 1):
        cell = new_ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = data_align
        cell.border = thin_border

    print(f"  [{i+1}] {position_name[:30]:30s} | {map_category(category_raw):10s} | {map_company_type(industry):20s} | {normalize_city(city_raw):15s} | {extract_project_year(category_raw, position_name)}")

output = "/Users/hw/workspace/OSGPrj/docs/岗位导入模板.xlsx"
new_wb.save(output)
print(f"\n已保存到: {output}")
print(f"共 {len(old_rows)} 条岗位数据已迁移")
