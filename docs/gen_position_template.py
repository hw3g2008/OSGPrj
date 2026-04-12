# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "岗位列表"

# 表头定义: (列名, 宽度, 是否必填, 说明)
headers = [
    ("岗位分类", 15, True, "如: summer / fulltime / offcycle / spring / events"),
    ("岗位名称", 25, True, "如: Summer Analyst / IB Analyst"),
    ("公司名称", 20, True, "如: Goldman Sachs / McKinsey"),
    ("公司类别", 18, True, "如: Investment Bank / Consulting / Tech / PE/VC / Other"),
    ("城市", 18, True, "如: New York / Hong Kong / Beijing"),
    ("招聘周期", 18, True, "如: 2025 Summer / Off-cycle / Rolling"),
    ("项目时间", 12, True, "如: 2025 / 2026 / 2027"),
    ("截止时间", 15, False, "日期格式或文字, 如: 2026-05-30 / Rolling ASAP"),
    ("部门", 20, False, "如: IBD / S&T（选填）"),
    ("岗位链接", 35, False, "投递链接 https://..."),
    ("公司官网", 35, False, "公司主页 https://..."),
]

# 样式
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
header_fill_req = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_fill_opt = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
wrap_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 写表头
for col_idx, (name, width, required, hint) in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=name + (" *" if required else ""))
    cell.font = header_font
    cell.fill = header_fill_req if required else header_fill_opt
    cell.alignment = wrap_align
    cell.border = thin_border
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

# 写说明行(第2行)
hint_font = Font(name="微软雅黑", size=9, color="666666", italic=True)
hint_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
for col_idx, (name, width, required, hint) in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_idx, value=hint)
    cell.font = hint_font
    cell.fill = hint_fill
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = thin_border

# 写示例数据
examples = [
    ["summer", "Summer Analyst", "Goldman Sachs", "Investment Bank", "New York", "2026 Summer", "2026", "2026-05-30", "IBD", "https://goldmansachs.com/careers/sa", "https://goldmansachs.com"],
    ["summer", "Summer Analyst", "McKinsey", "Consulting", "Beijing", "2026 Summer", "2026", "", "Consulting", "https://mckinsey.com/careers", "https://mckinsey.com"],
    ["fulltime", "Full-time Analyst", "Google", "Tech", "San Francisco", "2025 Full-time", "2025", "Rolling ASAP", "Engineering", "", "https://careers.google.com"],
]

data_font = Font(name="微软雅黑", size=10)
for row_idx, row_data in enumerate(examples, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border

# 冻结首行
ws.freeze_panes = "A3"
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 35

# 添加说明 sheet
ws2 = wb.create_sheet("填写说明")
instructions = [
    ["字段", "必填", "含义", "可选值"],
    ["岗位分类", "是", "岗位招聘类型", "summer(暑期实习) / fulltime(全职) / offcycle(非常规) / spring(春季) / events(活动)"],
    ["岗位名称", "是", "岗位全称", "自由填写"],
    ["公司名称", "是", "公司全称", "自由填写，支持搜索已有公司"],
    ["公司类别", "是", "公司所属行业类别", "Investment Bank / Consulting / Tech / PE/VC / PE / VC / Other"],
    ["城市", "是", "岗位所在城市（系统自动推断大区）", "New York / San Francisco / London / Hong Kong / Singapore / Beijing / Shanghai 等"],
    ["招聘周期", "是", "招聘批次", "2025 Summer / 2026 Summer / 2025 Full-time / Off-cycle 等"],
    ["项目时间", "是", "所属项目年份", "2024 / 2025 / 2026 / 2027"],
    ["截止时间", "否", "申请截止日期或文字说明", "日期如 2026-05-30，或文字如 Rolling ASAP"],
    ["部门", "否", "具体部门", "如 IBD / S&T / Consulting"],
    ["岗位链接", "否", "岗位投递链接", "https://..."],
    ["公司官网", "否", "公司招聘主页", "https://..."],
    ["", "", "", ""],
    ["时间字段说明", "", "", ""],
    ["项目时间", "", "招聘项目所属年份，如2025表示2025级招聘", ""],
    ["招聘周期", "", "具体招聘批次，如2026 Summer表示2026年暑期批次", ""],
    ["截止时间", "", "岗位申请截止日期，选填", ""],
    ["发布时间", "", "系统自动记录导入时间，无需填写", ""],
    ["展示时间", "", "系统自动设置（导入起90天），无需填写", ""],
]

header_font2 = Font(name="微软雅黑", size=11, bold=True)
for row_idx, row_data in enumerate(instructions, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1 or row_idx == 14:
            cell.font = header_font2
        else:
            cell.font = Font(name="微软雅黑", size=10)

ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 8
ws2.column_dimensions['C'].width = 40
ws2.column_dimensions['D'].width = 60

output_path = "/Users/hw/workspace/OSGPrj/docs/岗位导入模板.xlsx"
wb.save(output_path)
print(f"已生成: {output_path}")
